import openpyxl as xl
import os

current_dir = os.getcwd()
wb = xl.load_workbook(f'{current_dir}\database\数据库.xlsx')
name = []
emit = []
row_count = []
total_name = []
total_unit = []
total_usage = []
row_each = 0

sheets = wb.sheetnames  # 全部sheet，list格式
for sheet_name in sheets:
    sheet_read = wb[f'{sheet_name}']
    row_each = sheet_read.max_row + row_each
    for i in range(1, sheet_read.max_row + 1):
        each_name = sheet_read.cell(i, 1)
        each_unit = sheet_read.cell(i, 2)
        each_usage = sheet_read.cell(i, 3)
        total_name.append(each_name.value)
        total_unit.append(each_unit.value)
        total_usage.append(each_usage.value)
    row_count.append(row_each)
row_count = [0] + row_count
# print(row_count)  # 每个sheet多少行，list格式（含表头）
# print(len(total_name))
# print(total_name)  # 和rowcount对应的名称


def sheet(num):  # 输入0-5
    draw_name = []
    cor_unit = []
    cor_unit_emit = []
    if num == 0:
        pass
    else:
        for j in range(row_count[num - 1], row_count[num]):
            draw_name.append(total_name[j])  # 单个sheet第一列，取决于输入的num
            cor_unit.append(total_unit[j])
            cor_unit_emit.append(total_usage[j])

    return draw_name, cor_unit, cor_unit_emit
# 调用方法 sheet(a)[b], a 决定哪个sheet, 1混凝土 --- 5其他
# b决定哪一列,0名，1单位，2碳排
# sheet(a)[b][c], c能直接调用到列表的某一个item






