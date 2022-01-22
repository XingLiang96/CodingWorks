import sys
import datetime
import database
import openpyxl as xl
import os
from openpyxl import Workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *

wb = Workbook()  # 创建一个工作表
ws = wb.active   # ws操作sheet页
sheet = wb.create_sheet('Sheet1', 0)


class MyPyQT_Form(QtWidgets.QWidget):
    def __init__(self):
        super(MyPyQT_Form, self).__init__()
        self.setupUi(self)
        self.tableWidget.setEditTriggers(QTableView.NoEditTriggers)  # 不可编辑
        self.tableWidget.setColumnWidth(0, 210)
        self.tableWidget.setColumnWidth(1, 90)
        self.tableWidget.setColumnWidth(2, 90)
        self.tableWidget.setColumnWidth(3, 90)
        self.tableWidget.setColumnWidth(4, 50)
        self.loop_time = 0
        self.item_name = 0
        self.unit = 0
        self.unit_usage = 0
        self.usage_input = 0
        self.row = 0
        self.carbon = 0
        self.unit_each = 0
        self.usage_each = 0
        self.nested_usage = 0
        self.index = ['请选择材料种类', '混凝土', '水泥', '砂浆', '钢材', '其他']
        self.type_chosen = 0
        self.place = 0
        self.full_name = []
        self.length = []
        for big_type in self.index:  # 第一个下拉栏选项
            self.big_types.addItem(big_type)  # index里面的种类导入

    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.setEnabled(True)
        Form.resize(684, 892)
        font = QtGui.QFont()
        font.setPointSize(10)
        Form.setFont(font)
        self.label_6 = QtWidgets.QLabel(Form)
        self.label_6.setGeometry(QtCore.QRect(10, 860, 141, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.pushButton_2 = QtWidgets.QPushButton(Form)
        self.pushButton_2.setGeometry(QtCore.QRect(550, 840, 91, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(Form)
        self.pushButton_3.setGeometry(QtCore.QRect(80, 170, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(420, 220, 31, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.textEdit_2 = QtWidgets.QTextEdit(Form)
        self.textEdit_2.setEnabled(False)
        self.textEdit_2.setGeometry(QtCore.QRect(420, 240, 91, 31))
        self.textEdit_2.setObjectName("textEdit_2")
        self.label_4 = QtWidgets.QLabel(Form)
        self.label_4.setGeometry(QtCore.QRect(530, 220, 71, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.textEdit_3 = QtWidgets.QTextEdit(Form)
        self.textEdit_3.setEnabled(False)
        self.textEdit_3.setGeometry(QtCore.QRect(530, 240, 91, 31))
        self.textEdit_3.setObjectName("textEdit_3")
        self.label_5 = QtWidgets.QLabel(Form)
        self.label_5.setGeometry(QtCore.QRect(270, 220, 101, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.textEdit_4 = QtWidgets.QTextEdit(Form)
        self.textEdit_4.setEnabled(False)
        self.textEdit_4.setGeometry(QtCore.QRect(270, 240, 131, 31))
        self.textEdit_4.setObjectName("textEdit_4")
        self.pushButton_5 = QtWidgets.QPushButton(Form)
        self.pushButton_5.setGeometry(QtCore.QRect(80, 240, 101, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setObjectName("pushButton_5")
        self.tableWidget = QtWidgets.QTableWidget(Form)
        self.tableWidget.setEnabled(True)
        self.tableWidget.setGeometry(QtCore.QRect(50, 310, 570, 491))
        self.tableWidget.setMaximumSize(QtCore.QSize(570, 16777215))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.tableWidget.setFont(font)
        self.tableWidget.setRowCount(1)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(5)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(True)
        self.tableWidget.horizontalHeader().setSortIndicatorShown(False)
        self.tableWidget.verticalHeader().setVisible(True)
        self.big_types = QtWidgets.QComboBox(Form)
        self.big_types.setGeometry(QtCore.QRect(80, 50, 141, 31))
        self.big_types.setObjectName("big_types")
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(320, 110, 91, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.label_8 = QtWidgets.QLabel(Form)
        self.label_8.setGeometry(QtCore.QRect(200, 250, 61, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.pushButton_4 = QtWidgets.QPushButton(Form)
        self.pushButton_4.setGeometry(QtCore.QRect(540, 30, 81, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setObjectName("pushButton_4")
        self.sub_types = QtWidgets.QComboBox(Form)
        self.sub_types.setGeometry(QtCore.QRect(80, 110, 231, 31))
        self.sub_types.setObjectName("sub_types")
        self.label_9 = QtWidgets.QLabel(Form)
        self.label_9.setGeometry(QtCore.QRect(60, 30, 111, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(Form)
        self.label_10.setGeometry(QtCore.QRect(60, 90, 111, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setUnderline(True)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(Form)
        self.label_11.setGeometry(QtCore.QRect(60, 150, 131, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setUnderline(True)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(Form)
        self.label_12.setGeometry(QtCore.QRect(60, 220, 131, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setUnderline(True)
        self.label_12.setFont(font)
        self.label_12.setObjectName("label_12")
        self.pushButton_6 = QtWidgets.QPushButton(Form)
        self.pushButton_6.setGeometry(QtCore.QRect(450, 30, 81, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setObjectName("pushButton_6")

        self.retranslateUi(Form)
        self.pushButton_2.clicked.connect(Form.excel)
        self.pushButton_3.clicked.connect(Form.usage)
        self.pushButton_5.clicked.connect(Form.add)
        self.big_types.activated['QString'].connect(Form.types)
        self.pushButton.clicked.connect(Form.search)
        self.pushButton_4.clicked.connect(Form.open)
        self.sub_types.activated['QString'].connect(Form.draw)
        self.pushButton_6.clicked.connect(Form.manual)
        QtCore.QMetaObject.connectSlotsByName(Form)

        icon = QIcon('search.ico')
        self.pushButton.setIcon(icon)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "工程材料碳排放简易计算器V1.0"))
        self.label_6.setText(_translate("Form", "copyright @ 梁型"))
        self.pushButton_2.setText(_translate("Form", "生成统计表"))
        self.pushButton_3.setText(_translate("Form", "输入使用量"))
        self.label_3.setText(_translate("Form", "用量"))
        self.label_4.setText(_translate("Form", "碳排放"))
        self.label_5.setText(_translate("Form", "单位碳排放 "))
        self.pushButton_5.setText(_translate("Form", "记录该条目"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Form", "混凝土类型"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Form", "单位碳排放"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("Form", "用量"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("Form", "总碳排放"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("Form", "操作"))
        self.pushButton.setText(_translate("Form", "搜索材料"))
        self.label_8.setText(_translate("Form", "当前预览："))
        self.pushButton_4.setText(_translate("Form", "导入工作表"))
        self.label_9.setText(_translate("Form", "1.输入材料类别"))
        self.label_10.setText(_translate("Form", "2.选择具体材料"))
        self.label_11.setText(_translate("Form", "3.输入材料工程用量"))
        self.label_12.setText(_translate("Form", "4.预览并保存条目"))
        self.pushButton_6.setText(_translate("Form", "使用说明"))

    def types(self):  # 第一个下拉栏
        self.type_chosen = self.big_types.currentText()
        self.place = self.index.index(self.type_chosen)  # index是自定义的五个种类
        if self.place == 0:
            self.sub_types.clear()
            QMessageBox.about(self, "提示：", "请选择材料类别！")
            self.textEdit_4.clear()
        else:
            self.sub_types.clear()
            for name in database.sheet(self.place)[0]:
                self.sub_types.addItem(name)  # 添加小类
            self.textEdit_4.clear()

    def draw(self):  # 小类，第二个下拉栏
        self.item_name = self.sub_types.currentText()
        self.row = database.total_name.index(self.item_name)  # 定位所选的行数
        self.unit = database.total_unit[self.row]  # 对应单位
        self.unit_usage = database.total_usage[self.row]  # 对应碳排放
        self.textEdit_4.setText(f'{self.unit_usage}kg / {self.unit}')

    def usage(self):
        self.usage_input, ok = QInputDialog.getDouble(self, '用量', '输入用量（数字）：')  # 用量
        if self.usage_input == 0:
            QMessageBox.about(self, "提示：", "请输入非零的数字！")
        else:
            self.textEdit_2.setText(f'{self.usage_input} {self.unit}')
            self.usage_each = round(self.unit_usage * self.usage_input, 2)
            self.textEdit_3.setText(f'{self.usage_each} kg')

    def search(self):
        chars, ok = QInputDialog.getText(self, '搜索', '请输入搜索关键字：')
        i = 0
        self.sub_types.clear()
        for search_index in database.total_name:
            if chars in search_index:
                self.sub_types.addItem(search_index)
                i += 1
        if i == 0:
            QMessageBox.about(self, "提示：", "您输入的关键字不存在！")
        else:
            pass

    def add(self):
        # for shown list
        if self.usage_input == 0:
            QMessageBox.about(self, "提示：", "请输入材料使用量！")
        else:
            for i in range(0, self.loop_time + 1):
                item_name = QTableWidgetItem(str(self.item_name))
                self.tableWidget.setItem(self.loop_time, 0, item_name)
                carbon = QTableWidgetItem(str(self.unit_usage))
                self.tableWidget.setItem(self.loop_time, 1, carbon)
                concrete_usage = QTableWidgetItem(str(self.usage_input))
                self.tableWidget.setItem(self.loop_time, 2, concrete_usage)
                total_usage = QTableWidgetItem(str(self.usage_each))
                self.tableWidget.setItem(self.loop_time, 3, total_usage)

            self.textEdit_2.setText("/")
            self.textEdit_3.setText("/")
            self.textEdit_4.setText("/")
            self.loop_time += 1   # 初始是0
            self.nested_usage = round(self.nested_usage + self.usage_each, 2)
            self.usage_input = 0
            for i in range(self.loop_time):
                self.tableWidget.setCellWidget(i, 4, self.buttonForRow())  # 在最后一个单元格中加入修改、删除按钮
            self.tableWidget.setRowCount(self.loop_time + 1)
            QMessageBox.about(self, "提示：", "记录完成,请输入新条目！")

    def buttonForRow(self):
        widget = QtWidgets.QWidget()

        # 删除
        self.deleteBtn = QtWidgets.QPushButton('删除')
        self.deleteBtn.setStyleSheet(''' text-align : center;
                                       background-color : LightCoral;
                                       height : 30px;
                                       border-style: outset;
                                       font : 13px; ''')
        self.deleteBtn.clicked.connect(self.DeleteButton)
        hLayout = QtWidgets.QHBoxLayout()
        hLayout.addWidget(self.deleteBtn)
        hLayout.setContentsMargins(2, 2, 5, 2)
        widget.setLayout(hLayout)
        return widget

    def DeleteButton(self):
        reply = QMessageBox.question(self, '删除', '你确定要删除吗?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            button = self.sender()
            if button:
                #  确定位置的时候这里是关键
                row = self.tableWidget.indexAt(button.parent().pos()).row()
                self.tableWidget.removeRow(row)
                # print(row)
                self.loop_time -= 1
        else:
            pass

    def excel(self):
        total_carbon = 0
        name = []
        unit = []
        use = []
        total = []

        for i in range(self.loop_time):  # tableWidget的内容导入
            col1 = sheet.cell(i + 3, 1)  # 序号
            col1.value = i + 1

            a = self.tableWidget.item(i, 0).text()  # 读取tableWidget
            name.append(a)
            col2 = sheet.cell(i + 3, 2)  # 存入xl
            col2.value = name[i]

            b = self.tableWidget.item(i, 1).text()
            unit.append(b)
            col3 = sheet.cell(i + 3, 3)
            col3.value = unit[i]

            c = self.tableWidget.item(i, 2).text()
            use.append(c)
            col4 = sheet.cell(i + 3, 4)
            col4.value = use[i]

            d = self.tableWidget.item(i, 3).text()
            total.append(d)
            col5 = sheet.cell(i + 3, 5)
            col5.value = total[i]

        for each in total:  # 总碳排放计算
            num = float(each)
            total_carbon += num

        # 设置列宽
        sheet.column_dimensions['a'].width = 15
        sheet.column_dimensions['b'].width = 30
        sheet.column_dimensions['c'].width = 15
        sheet.column_dimensions['d'].width = 15
        sheet.column_dimensions['e'].width = 15

        # title
        f1 = sheet.cell(1, 1)
        f1.value = "工作名："
        h1 = sheet.cell(1, 3)
        h1.value = "日期："
        a1 = sheet.cell(2, 1)
        a1.value = "序号"
        b1 = sheet.cell(2, 2)
        b1.value = "使用材料"
        c1 = sheet.cell(2, 3)
        c1.value = "单位碳排放"
        d1 = sheet.cell(2, 4)
        d1.value = "用量"
        e1 = sheet.cell(2, 5)
        e1.value = "总碳排放（kg）"

        con_title = sheet.cell(self.loop_time + 5, 1)
        con_title.value = f"二氧化碳总碳排放量：{round(total_carbon, 2)} kg"
        file_name, ok = QInputDialog.getText(self, '文件名', '请命名文件：')
        curr_time = datetime.datetime.now()
        time_str = curr_time.strftime("%Y%m%d%H%M")

        g1 = sheet.cell(1, 2)
        g1.value = f"{file_name}"
        i1 = sheet.cell(1, 4)
        i1.value = f"{time_str}"

        wb.save(f'{time_str}-{file_name}.xlsx')
        QMessageBox.about(self, "提示：", f"保存完成！文件名为“{time_str}{file_name}统计表.xlsx”")

    def open(self):
        QMessageBox.about(self, "提示：", "导入操作会覆盖当前所有数据。请确保所有数据已经保存！")
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        path_openfile_name = openfile_name[0]
        if len(path_openfile_name) > 0:
            self.tableWidget.setRowCount(0)
            # print(path_openfile_name)
            wb = xl.load_workbook(path_openfile_name)
            sheet = wb["Sheet1"]
            rowcount = sheet.max_row - 5
            self.tableWidget.setRowCount(rowcount + 1)
            self.loop_time = rowcount + 1
            for row in range(2, sheet.max_row - 3):  # sheet.max_row/col --- will be an interger
                # print(sheet.max_row)
                for col in range(2, 6):
                    item = sheet.cell(row + 1, col)
                    item2list = item.value
                    # print(item2list)
                    table_item = QTableWidgetItem(str(item2list))
                    self.tableWidget.setItem(row - 2, col - 2, table_item)

            for i in range(self.loop_time - 1):
                self.tableWidget.setCellWidget(i, 4, self.buttonForRow())  # 在最后一个单元格中加入修改、删除按钮

            QMessageBox.about(self, "提示：", "导入完成！")
            self.loop_time = rowcount
        else:
            pass

    def manual(self):
        QMessageBox.about(self, "使用说明：", "-------------------------------------使用流程--------------------------------------\n"
                                         "1. 在‘输入材料类别’里下拉选择材料的类别；\n"
                                         "2. 选择完材料类别后，在第二个下拉栏里选择具体的某种材料；\n"
                                         "3. 输入材料的使用量，只能为数字；\n"
                                         "4. 在‘当前预览’可以预览你选择的材料和使用量，确认无误后点击记录即可保存条目；\n"
                                         "5. 全部条目添加完成后，点击右下方的'生成统计表'按钮即可生成excel文件\n"
                                         "6. 如记录的条目有误，点击条目后方的‘删除’即可。\n"
                                         "----------------------------------执行现有文件-----------------------------------\n"
                                         "通过上述流程步骤生成的excel文件，可以导入到该软件继续工作\n"
                                         "1. 点击右上方‘导入工作表’按钮，选择文件\n"
                                         "2. 重复上面使用流程的1~6步即可对已有文件进行增、删、查、改等操作；\n"
                                         "3. 操作完成后，点生成统计表，才会生成新的文件。（不会改写原文件）")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    demo = MyPyQT_Form()
    demo.show()
    sys.exit(app.exec_())